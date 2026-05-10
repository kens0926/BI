"""
控制點資料模組 (SQLite)
"""
from dataclasses import dataclass, field
from typing import List, Dict, Optional
from datetime import datetime, timedelta
from database import get_db_connection


# 控制點分類
CONTROL_CATEGORIES = [
    "存取控制",
    "帳務核覆",
    "資安管理",
    "採購流程",
    "人事管理",
    "財務管理",
    "風險管理",
    "法規遵循",
]

# 業務流程
PROCESSES = [
    "財務管理",
    "資訊安全",
    "採購管理",
    "人力資源",
    "生產製造",
    "業務銷售",
    "法務合規",
    "品質管理",
]

# 風險等級
RISK_LEVELS = {"H": "高", "M": "中", "L": "低"}

# 測試頻率
TEST_FREQUENCIES = ["每月", "每季", "每半年", "每年"]

# 測試結果
TEST_RESULTS = ["通過", "不通過", "不適用", "待測試"]

# 控制點狀態
CONTROL_STATUS = ["有效", "停用", "審查中"]

# 控制測試流程狀態 (6.2)
TEST_WORKFLOW_STATES = ["排程建立", "待執行", "測試中", "待覆核", "已完成"]

# 控制類型 (CTL-04)
CONTROL_TYPES = ["預防性控制", "偵測性控制", "更正性控制"]

# 控制模式
CONTROL_MODES = ["人工", "半自動", "全自動"]

# 測試方法
TEST_METHODS = ["訪談", "文件檢查", "重新執行", "資料分析"]

# 抽樣方法
SAMPLE_METHODS = ["判斷性抽樣", "隨機抽樣", "全查"]


@dataclass
class ControlTest:
    """控制點測試記錄 (對應 7.4.4)"""
    id: int
    control_id: str
    test_date: str
    test_result: str
    tester: str
    # 新增欄位
    test_period: str = ""  # 測試期間
    test_method: str = ""  # 測試方法
    sample_method: str = ""  # 抽樣方法
    sample_size: int = 0  # 抽樣件數
    exception_count: int = 0  # 異常件數
    reviewer: str = ""  # 覆核人
    findings: str = ""
    evidence_files: List[str] = field(default_factory=list)
    related_issue_id: Optional[int] = None
    # 新增流程狀態
    test_status: str = "待執行"  # 排程建立/待執行/測試中/待覆核/已完成


@dataclass
class ControlPoint:
    """控制點資料結構 (對應 7.4.3)"""
    id: str
    name: str
    category: str
    process: str
    risk_level: str
    description: str
    test_frequency: str
    # 新增欄位對應 7.4.3
    control_objective: str = ""  # 控制目標
    control_type: str = ""  # 預防性/偵測性/更正性
    control_mode: str = ""  # 人工/半自動/全自動
    key_control_flag: bool = False  # 是否為關鍵控制
    effective_from: str = ""  # 生效日
    effective_to: str = ""  # 失效日
    version_no: str = "V1.0"  # 版本號
    last_test_date: str = ""
    last_test_result: str = "待測試"
    next_test_date: str = ""
    owner_dept: str = ""
    owner_user: str = ""
    related_issues: List[int] = field(default_factory=list)
    evidence_files: List[str] = field(default_factory=list)
    status: str = "有效"
    created_at: str = ""
    updated_at: str = ""
    test_status: str = "排程建立"  # 測試流程狀態

    def calculate_next_test_date(self) -> str:
        """根據測試頻率自動計算下次測試日期"""
        if not self.last_test_date:
            return ""
        
        try:
            last_date = datetime.strptime(self.last_test_date, "%Y-%m-%d")
        except:
            return ""
        
        freq_map = {
            "每月": 30,
            "每季": 90,
            "每半年": 180,
            "每年": 365,
        }
        
        days = freq_map.get(self.test_frequency, 30)
        next_date = last_date + timedelta(days=days)
        return next_date.strftime("%Y-%m-%d")

    def is_overdue(self) -> bool:
        """檢查是否逾期"""
        if not self.next_test_date:
            return False
        try:
            next_date = datetime.strptime(self.next_test_date, "%Y-%m-%d")
            return next_date.date() < datetime.now().date()
        except:
            return False

    def days_until_test(self) -> int:
        """距離下次測試的天數"""
        if not self.next_test_date:
            return 999
        try:
            next_date = datetime.strptime(self.next_test_date, "%Y-%m-%d")
            delta = next_date - datetime.now()
            return delta.days
        except:
            return 999

    def next_test_statuses(self) -> List[str]:
        """取得可轉換的下一個測試狀態"""
        flow = {
            "排程建立": ["待執行"],
            "待執行": ["測試中"],
            "測試中": ["待覆核"],
            "待覆核": ["已完成"],
            "已完成": [],
        }
        return flow.get(self.test_status, [])

    def can_transition_to(self, new_status: str) -> bool:
        """檢查是否可以轉換到指定狀態"""
        return new_status in self.next_test_statuses()


class ControlManager:
    """控制點管理器"""
    
    def __init__(self):
        self._load_controls()
        self._load_tests()

    def _load_controls(self):
        """從資料庫載入控制點"""
        self._controls = {}
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM controls ORDER BY id")
        rows = cursor.fetchall()
        conn.close()
        
        for row in rows:
            self._controls[row["id"]] = self._row_to_control(row)

    def _load_tests(self):
        """從資料庫載入測試記錄"""
        self._tests = {}
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM control_tests ORDER BY id")
        rows = cursor.fetchall()
        conn.close()
        
        for row in rows:
            self._tests[row["id"]] = self._row_to_test(row)
        
        self._next_test_id = max(self._tests.keys(), default=0) + 1

    def list_controls(self) -> List[ControlPoint]:
        """取得所有控制點清單"""
        return sorted(self._controls.values(), key=lambda c: c.id)

    def get_control(self, control_id: str) -> Optional[ControlPoint]:
        """取得指定控制點"""
        return self._controls.get(control_id)

    def _row_to_control(self, row) -> ControlPoint:
        """將資料庫資料轉換為 ControlPoint 物件"""
        import json
        related = []
        evidence = []
        try:
            related = json.loads(row["related_issues"]) if row["related_issues"] else []
        except:
            pass
        try:
            evidence = json.loads(row["evidence_files"]) if row["evidence_files"] else []
        except:
            pass
        
        return ControlPoint(
            id=row["id"],
            name=row["name"],
            category=row["category"] or "",
            process=row["process"] or "",
            risk_level=row["risk_level"] or "M",
            description=row["description"] or "",
            test_frequency=row["test_frequency"] or "每月",
            control_objective=row["control_objective"] or "",
            control_type=row["control_type"] or "",
            control_mode=row["control_mode"] or "",
            key_control_flag=bool(row["key_control_flag"] or 0),
            effective_from=row["effective_from"] or "",
            effective_to=row["effective_to"] or "",
            version_no=row["version_no"] or "V1.0",
            last_test_date=row["last_test_date"] or "",
            last_test_result=row["last_test_result"] or "待測試",
            next_test_date=row["next_test_date"] or "",
            owner_dept=row["owner_dept"] or "",
            owner_user=row["owner_user"] or "",
            related_issues=related,
            evidence_files=evidence,
            status=row["status"] or "有效",
            created_at=row["created_at"] or "",
            updated_at=row["updated_at"] or "",
            test_status=row["test_status"] if "test_status" in row.keys() else "排程建立",
        )

    def _row_to_test(self, row) -> ControlTest:
        """將資料庫資料轉換為 ControlTest 物件"""
        import json
        evidence = []
        try:
            evidence = json.loads(row["evidence_files"]) if row["evidence_files"] else []
        except:
            pass
        
        return ControlTest(
            id=row["id"],
            control_id=row["control_id"],
            test_date=row["test_date"] or "",
            test_result=row["test_result"] or "",
            tester=row["tester"] or "",
            test_period=row["test_period"] or "",
            test_method=row["test_method"] or "",
            sample_method=row["sample_method"] or "",
            sample_size=row["sample_size"] or 0,
            exception_count=row["exception_count"] or 0,
            reviewer=row["reviewer"] or "",
            findings=row["findings"] or "",
            evidence_files=evidence,
            related_issue_id=row["related_issue_id"],
            test_status=row["test_status"] or "待執行",
        )

    # ===== CRUD 操作 =====

    def create_control(
        self,
        id: str,
        name: str,
        category: str,
        process: str,
        risk_level: str,
        description: str,
        test_frequency: str,
        owner_dept: str,
        owner_user: str,
    ) -> ControlPoint:
        """新增控制點"""
        import json
        
        if id in self._controls:
            raise ValueError(f"控制點 ID {id} 已存在")
        
        now = datetime.now().strftime("%Y-%m-%d")
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            INSERT INTO controls (id, name, category, process, risk_level, 
                description, test_frequency, owner_dept, owner_user, status, 
                created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, '有效', ?, ?)
        """, (id, name, category, process, risk_level, description, 
              test_frequency, owner_dept, owner_user, now, now))
        
        conn.commit()
        conn.close()
        
        self._load_controls()
        return self._controls.get(id)

    def update_control(
        self,
        control_id: str,
        name: str = None,
        category: str = None,
        process: str = None,
        risk_level: str = None,
        description: str = None,
        test_frequency: str = None,
        owner_dept: str = None,
        owner_user: str = None,
        status: str = None,
        test_status: str = None,
    ) -> Optional[ControlPoint]:
        """更新控制點"""
        control = self._controls.get(control_id)
        if not control:
            return None
        
        # Build update fields
        fields = {}
        if name is not None:
            fields["name"] = name
        if category is not None:
            fields["category"] = category
        if process is not None:
            fields["process"] = process
        if risk_level is not None:
            fields["risk_level"] = risk_level
        if description is not None:
            fields["description"] = description
        if test_frequency is not None:
            fields["test_frequency"] = test_frequency
        if owner_dept is not None:
            fields["owner_dept"] = owner_dept
        if owner_user is not None:
            fields["owner_user"] = owner_user
        if status is not None:
            fields["status"] = status
        if test_status is not None:
            fields["test_status"] = test_status
        
        if fields:
            fields["updated_at"] = datetime.now().strftime("%Y-%m-%d")
            conn = get_db_connection()
            cursor = conn.cursor()
            
            set_clause = ", ".join([f"{k} = ?" for k in fields.keys()])
            values = list(fields.values()) + [control_id]
            
            cursor.execute(f"UPDATE controls SET {set_clause} WHERE id = ?", values)
            conn.commit()
            conn.close()
        
        self._load_controls()
        return self._controls.get(control_id)

    def delete_control(self, control_id: str) -> bool:
        """停用控制點 (軟刪除)"""
        control = self._controls.get(control_id)
        if not control:
            return False

        now = datetime.now().strftime("%Y-%m-%d")
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE controls SET status = ?, updated_at = ? WHERE id = ?",
            ("停用", now, control_id)
        )
        conn.commit()
        conn.close()

        # 同步記憶體狀態
        control.status = "停用"
        control.updated_at = now
        return True

    # ===== 測試記錄操作 =====

    def record_test(
        self,
        control_id: str,
        test_date: str,
        test_result: str,
        tester: str,
        findings: str = "",
        evidence_files: List[str] = None,
        related_issue_id: int = None,
    ) -> Optional[ControlTest]:
        """登錄測試結果"""
        control = self._controls.get(control_id)
        if not control:
            return None
        
        test = ControlTest(
            id=self._next_test_id,
            control_id=control_id,
            test_date=test_date,
            test_result=test_result,
            tester=tester,
            findings=findings,
            evidence_files=evidence_files or [],
            related_issue_id=related_issue_id,
        )
        
        self._tests[self._next_test_id] = test
        self._next_test_id += 1
        
        # 更新控制點的最後測試資訊
        control.last_test_date = test_date
        control.last_test_result = test_result
        control.next_test_date = control.calculate_next_test_date()
        control.updated_at = datetime.now().strftime("%Y-%m-%d")
        
        # 如果發現問題且有關聯 Issue
        if test_result == "不通過" and related_issue_id:
            if related_issue_id not in control.related_issues:
                control.related_issues.append(related_issue_id)
        
        return test

    def get_control_tests(self, control_id: str) -> List[ControlTest]:
        """取得控制點的測試歷史"""
        return [t for t in self._tests.values() if t.control_id == control_id]

    # ===== 統計與分析 =====

    def get_coverage_stats(self) -> Dict:
        """取得控制覆蓋率統計"""
        total = len(self._controls)
        active = sum(1 for c in self._controls.values() if c.status == "有效")
        
        # 按流程統計
        by_process = {}
        for c in self._controls.values():
            if c.status == "有效":
                if c.process not in by_process:
                    by_process[c.process] = {"total": 0, "active": 0}
                by_process[c.process]["total"] += 1
                by_process[c.process]["active"] += 1
        
        # 按風險等級統計
        by_risk = {"H": 0, "M": 0, "L": 0}
        for c in self._controls.values():
            if c.status == "有效" and c.risk_level in by_risk:
                by_risk[c.risk_level] += 1
        
        return {
            "total": total,
            "active": active,
            "coverage": round(active / total * 100, 1) if total > 0 else 0,
            "by_process": by_process,
            "by_risk": by_risk,
        }

    def get_upcoming_tests(self, days: int = 30) -> List[ControlPoint]:
        """取得即將到期的測試"""
        upcoming = []
        for control in self._controls.values():
            if control.status == "有效" and control.next_test_date:
                if control.days_until_test() <= days:
                    upcoming.append(control)
        return sorted(upcoming, key=lambda x: x.days_until_test())

    def get_overdue_controls(self) -> List[ControlPoint]:
        """取得逾期未測試的控制點"""
        return [c for c in self._controls.values() if c.status == "有效" and c.is_overdue()]

    def search_controls(
        self,
        process: str = None,
        category: str = None,
        risk_level: str = None,
        status: str = None,
        keyword: str = None,
    ) -> List[ControlPoint]:
        """搜尋控制點"""
        results = self._controls.values()
        
        if process:
            results = [c for c in results if c.process == process]
        if category:
            results = [c for c in results if c.category == category]
        if risk_level:
            results = [c for c in results if c.risk_level == risk_level]
        if status:
            results = [c for c in results if c.status == status]
        if keyword:
            keyword_lower = keyword.lower()
            results = [
                c for c in results
                if keyword_lower in c.name.lower()
                or keyword_lower in c.description.lower()
                or keyword_lower in c.id.lower()
            ]
        
        return sorted(results, key=lambda x: x.id)

    def import_from_excel(self, file_path: str) -> Dict[str, any]:
        """從Excel文件匯入控制點"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {"success": False, "message": "需要安裝 openpyxl 套件"}
        
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            imported = 0
            errors = []
            
            # 跳過header行
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row[0]:  # 如果ID為空，跳過
                        continue
                    
                    # 提取欄位
                    control_id = str(row[0]).strip()
                    name = str(row[1]).strip() if row[1] else ""
                    category = str(row[2]).strip() if row[2] else ""
                    process = str(row[3]).strip() if row[3] else ""
                    risk_level = str(row[4]).strip() if row[4] else "M"
                    description = str(row[5]).strip() if row[5] else ""
                    test_frequency = str(row[6]).strip() if row[6] else "每季"
                    
                    # 驗證必填欄位
                    if not control_id or not name:
                        errors.append(f"第 {row_idx} 行：控制點ID和名稱為必填")
                        continue
                    
                    # 檢查是否已存在
                    if control_id in self._controls:
                        # 更新現有控制點
                        self.update_control(
                            control_id, name=name, category=category, 
                            process=process, risk_level=risk_level,
                            description=description, test_frequency=test_frequency
                        )
                    else:
                        # 新增控制點
                        self.create_control(
                            control_id, name, category, process, 
                            risk_level, description, test_frequency,
                            owner_dept="", owner_user=""
                        )
                    
                    imported += 1
                    
                except Exception as e:
                    errors.append(f"第 {row_idx} 行：{str(e)}")
            
            return {
                "success": True,
                "imported": imported,
                "errors": errors,
                "message": f"成功匯入 {imported} 筆控制點"
            }
            
        except Exception as e:
            return {"success": False, "message": f"匯入失敗：{str(e)}"}

    def export_to_excel(self, file_path: str, controls: List[ControlPoint] = None) -> bool:
        """匯出控制點到Excel文件"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return False
        
        try:
            if controls is None:
                controls = self.list_controls()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "控制點"
            
            # 設置header
            headers = ["控制點ID", "控制點名稱", "分類", "業務流程", "風險等級", "描述", "測試頻率"]
            ws.append(headers)
            
            # 設置header樣式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 添加資料
            for ctrl in controls:
                ws.append([
                    ctrl.id,
                    ctrl.name,
                    ctrl.category,
                    ctrl.process,
                    RISK_LEVELS.get(ctrl.risk_level, ctrl.risk_level),
                    ctrl.description,
                    ctrl.test_frequency
                ])
            
            # 調整列寬
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 30
            ws.column_dimensions['G'].width = 12
            
            wb.save(file_path)
            return True
            
        except Exception as e:
            print(f"匯出失敗：{str(e)}")
            return False


# 全域控制點管理器實例
control_manager = ControlManager()
