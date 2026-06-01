from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from ..database import get_db
from ..deps import current_user
from ..models import AuditPlan, AuditQuestion, AuditRecord, AuditResultType, User
from ..schemas import ApiResponse, AuditRecordCreate, AuditRecordRead, AuditRecordUpdate

router = APIRouter(prefix="/api/audit-records", tags=["audit-records"])

EXCEL_COLUMNS = [
    ("id", "記錄ID"),
    ("question_id", "題目ID"),
    ("question_text", "題目內容"),
    ("result_type", "查核結果"),
    ("finding", "查核說明"),
    ("suggestion", "改善建議"),
    ("attachment_path", "附件路徑"),
]

EXCEL_HEADER_ALIASES = {key: {key, label} for key, label in EXCEL_COLUMNS}


def validate_template(header: list[str]) -> None:
    expected = [label for _, label in EXCEL_COLUMNS]
    keys = [key for key, _ in EXCEL_COLUMNS]
    for index, key in enumerate(keys):
        value = header[index] if index < len(header) else ""
        if value not in EXCEL_HEADER_ALIASES[key]:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Invalid template. Expected columns: {', '.join(expected)}. "
                    f"Got: {', '.join(header)}"
                ),
            )


def parse_result_type(value) -> AuditResultType:
    if value in (None, ""):
        return AuditResultType.pass_
    normalized = str(value).strip().upper()
    for result_type in AuditResultType:
        if normalized == result_type.value:
            return result_type
    raise HTTPException(status_code=400, detail=f"Invalid audit result type: {value}")


@router.get("", response_model=ApiResponse)
def list_records(
    audit_plan_id: str | None = None,
    result_type: str | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(current_user),
):
    query = db.query(AuditRecord)
    if audit_plan_id:
        query = query.filter(AuditRecord.audit_plan_id == audit_plan_id)
    if result_type:
        query = query.filter(AuditRecord.result_type == result_type)
    rows = query.order_by(AuditRecord.created_at.desc()).all()
    return ApiResponse(message="Data retrieved successfully", data=[AuditRecordRead.model_validate(row) for row in rows])


@router.get("/export")
def export_records(
    audit_plan_id: str,
    db: Session = Depends(get_db),
    _: User = Depends(current_user),
):
    plan = db.get(AuditPlan, audit_plan_id)
    if not plan:
        raise HTTPException(status_code=404, detail="Audit plan not found")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "audit_records"
    sheet.append([label for _, label in EXCEL_COLUMNS])

    rows = (
        db.query(AuditRecord)
        .filter(AuditRecord.audit_plan_id == audit_plan_id)
        .order_by(AuditRecord.created_at.desc())
        .all()
    )
    for row in rows:
        sheet.append(
            [
                row.id,
                row.question_id,
                row.question.question if row.question else "",
                row.result_type.value,
                row.finding or "",
                row.suggestion or "",
                row.attachment_path or "",
            ]
        )

    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells) + 4
        sheet.column_dimensions[column_cells[0].column_letter].width = min(width, 60)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"audit_records_{plan.task_no}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/import", response_model=ApiResponse)
async def import_records(
    audit_plan_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(current_user),
):
    if not db.get(AuditPlan, audit_plan_id):
        raise HTTPException(status_code=404, detail="Audit plan not found")
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")

    content = await file.read()
    workbook = load_workbook(BytesIO(content), data_only=True)
    sheet = workbook.active
    header = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    validate_template(header)

    keys = [key for key, _ in EXCEL_COLUMNS]
    created_count = 0
    updated_count = 0
    for row_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(values):
            continue
        data = dict(zip(keys, values))
        question_id = str(data.get("question_id") or "").strip()
        if not question_id:
            raise HTTPException(status_code=400, detail=f"Row {row_index}: question_id is required")
        if not db.get(AuditQuestion, question_id):
            raise HTTPException(status_code=400, detail=f"Row {row_index}: question not found")

        payload = {
            "audit_plan_id": audit_plan_id,
            "question_id": question_id,
            "result_type": parse_result_type(data.get("result_type")),
            "finding": str(data.get("finding") or "").strip() or None,
            "suggestion": str(data.get("suggestion") or "").strip() or None,
            "attachment_path": str(data.get("attachment_path") or "").strip() or None,
        }
        record_id = str(data.get("id") or "").strip()
        row = db.get(AuditRecord, record_id) if record_id else None
        if row:
            for key, value in payload.items():
                setattr(row, key, value)
            updated_count += 1
        else:
            if record_id:
                payload["id"] = record_id
            db.add(AuditRecord(**payload, created_by=user.id))
            created_count += 1

    db.commit()
    return ApiResponse(
        message="Audit records imported",
        data={"created": created_count, "updated": updated_count},
    )


@router.post("", response_model=ApiResponse)
def create_record(payload: AuditRecordCreate, db: Session = Depends(get_db), user: User = Depends(current_user)):
    if not db.get(AuditPlan, payload.audit_plan_id):
        raise HTTPException(status_code=404, detail="Audit plan not found")
    if not db.get(AuditQuestion, payload.question_id):
        raise HTTPException(status_code=404, detail="Question not found")
    row = AuditRecord(**payload.model_dump(), created_by=user.id)
    db.add(row)
    db.commit()
    db.refresh(row)
    return ApiResponse(message="Audit record created", data=AuditRecordRead.model_validate(row))


@router.put("/{record_id}", response_model=ApiResponse)
def update_record(
    record_id: str,
    payload: AuditRecordUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(current_user),
):
    row = db.get(AuditRecord, record_id)
    if not row:
        raise HTTPException(status_code=404, detail="Audit record not found")

    update_data = payload.model_dump(exclude_unset=True)
    if "audit_plan_id" in update_data and not db.get(AuditPlan, update_data["audit_plan_id"]):
        raise HTTPException(status_code=404, detail="Audit plan not found")
    if "question_id" in update_data and not db.get(AuditQuestion, update_data["question_id"]):
        raise HTTPException(status_code=404, detail="Question not found")

    for key, value in update_data.items():
        setattr(row, key, value)
    db.commit()
    db.refresh(row)
    return ApiResponse(message="Audit record updated", data=AuditRecordRead.model_validate(row))
