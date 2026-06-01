from datetime import date, datetime
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from ..database import get_db
from ..deps import current_user
from ..models import AuditPlan, AuditPlanStatus, User
from ..schemas import ApiResponse, AuditPlanCreate, AuditPlanRead, AuditPlanUpdate

router = APIRouter(prefix="/api/audit-plans", tags=["audit-plans"])

EXCEL_COLUMNS = [
    ("task_no", "任務編號"),
    ("year", "稽核年度"),
    ("cycle_name", "稽核循環"),
    ("department", "被稽核部門"),
    ("auditor_name", "稽核員"),
    ("start_date", "開始日期"),
    ("end_date", "結束日期"),
    ("status", "任務狀態"),
]

EXCEL_HEADER_ALIASES = {
    key: {key, label}
    for key, label in EXCEL_COLUMNS
}


def parse_excel_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            pass
    raise HTTPException(status_code=400, detail=f"Invalid date value: {value}")


def parse_status(value) -> AuditPlanStatus:
    if value in (None, ""):
        return AuditPlanStatus.draft
    normalized = str(value).strip()
    for status in AuditPlanStatus:
        if normalized.lower() == status.value.lower():
            return status
    raise HTTPException(status_code=400, detail=f"Invalid task status: {value}")


@router.get("", response_model=ApiResponse)
def list_audit_plans(
    year: int | None = None,
    status: str | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(current_user),
):
    query = db.query(AuditPlan)
    if year:
        query = query.filter(AuditPlan.year == year)
    if status:
        query = query.filter(AuditPlan.status == status)
    rows = query.order_by(AuditPlan.year.desc(), AuditPlan.created_at.desc()).all()
    return ApiResponse(message="Data retrieved successfully", data=[AuditPlanRead.model_validate(row) for row in rows])


@router.get("/export")
def export_audit_plans(db: Session = Depends(get_db), _: User = Depends(current_user)):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "audit_plans"
    sheet.append([label for _, label in EXCEL_COLUMNS])

    rows = db.query(AuditPlan).order_by(AuditPlan.year.desc(), AuditPlan.task_no.asc()).all()
    for row in rows:
        sheet.append(
            [
                row.task_no,
                row.year,
                row.cycle_name,
                row.department,
                row.auditor_name or "",
                row.start_date,
                row.end_date,
                row.status.value,
            ]
        )

    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells) + 4
        sheet.column_dimensions[column_cells[0].column_letter].width = min(width, 30)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"audit_plans_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/import", response_model=ApiResponse)
async def import_audit_plans(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(current_user),
):
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")

    content = await file.read()
    workbook = load_workbook(BytesIO(content), data_only=True)
    sheet = workbook.active
    header = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    expected = [label for _, label in EXCEL_COLUMNS]
    keys = [key for key, _ in EXCEL_COLUMNS]
    for index, key in enumerate(keys):
        value = header[index] if index < len(header) else ""
        if value not in EXCEL_HEADER_ALIASES[key]:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Invalid template. Expected columns: {', '.join(expected)}. "
                    f"Got: {', '.join(header)}"
                ),
            )

    created_count = 0
    updated_count = 0
    for row_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(values):
            continue
        data = dict(zip([key for key, _ in EXCEL_COLUMNS], values))
        if not data.get("task_no"):
            raise HTTPException(status_code=400, detail=f"Row {row_index}: task number is required")
        if not data.get("year"):
            raise HTTPException(status_code=400, detail=f"Row {row_index}: audit year is required")

        row = db.query(AuditPlan).filter(AuditPlan.task_no == str(data["task_no"]).strip()).first()
        payload = {
            "task_no": str(data["task_no"]).strip(),
            "year": int(data["year"]),
            "cycle_name": str(data.get("cycle_name") or "").strip(),
            "department": str(data.get("department") or "").strip(),
            "auditor_name": str(data.get("auditor_name") or "").strip() or None,
            "start_date": parse_excel_date(data.get("start_date")),
            "end_date": parse_excel_date(data.get("end_date")),
            "status": parse_status(data.get("status")),
        }
        if not payload["cycle_name"] or not payload["department"]:
            raise HTTPException(status_code=400, detail=f"Row {row_index}: cycle and department are required")

        if row:
            for key, value in payload.items():
                setattr(row, key, value)
            updated_count += 1
        else:
            db.add(AuditPlan(**payload))
            created_count += 1

    db.commit()
    return ApiResponse(
        message="Audit plans imported",
        data={"created": created_count, "updated": updated_count},
    )


@router.post("", response_model=ApiResponse)
def create_audit_plan(payload: AuditPlanCreate, db: Session = Depends(get_db), _: User = Depends(current_user)):
    if db.query(AuditPlan).filter(AuditPlan.task_no == payload.task_no).first():
        raise HTTPException(status_code=409, detail="Task number already exists")
    row = AuditPlan(**payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return ApiResponse(message="Audit plan created", data=AuditPlanRead.model_validate(row))


@router.put("/{plan_id}", response_model=ApiResponse)
def update_audit_plan(
    plan_id: str,
    payload: AuditPlanUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(current_user),
):
    row = db.get(AuditPlan, plan_id)
    if not row:
        raise HTTPException(status_code=404, detail="Audit plan not found")
    update_data = payload.model_dump(exclude_unset=True)
    if "task_no" in update_data:
        duplicate = db.query(AuditPlan).filter(AuditPlan.task_no == update_data["task_no"], AuditPlan.id != plan_id).first()
        if duplicate:
            raise HTTPException(status_code=409, detail="Task number already exists")
    for key, value in update_data.items():
        setattr(row, key, value)
    db.commit()
    db.refresh(row)
    return ApiResponse(message="Audit plan updated", data=AuditPlanRead.model_validate(row))


@router.delete("/{plan_id}", response_model=ApiResponse)
def delete_audit_plan(plan_id: str, db: Session = Depends(get_db), _: User = Depends(current_user)):
    row = db.get(AuditPlan, plan_id)
    if not row:
        raise HTTPException(status_code=404, detail="Audit plan not found")
    db.delete(row)
    db.commit()
    return ApiResponse(message="Audit plan deleted", data={"id": plan_id})
