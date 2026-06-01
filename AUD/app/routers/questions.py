from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import or_
from sqlalchemy.orm import Session

from ..database import get_db
from ..deps import current_user
from ..models import AuditQuestion, User
from ..schemas import ApiResponse, AuditQuestionCreate, AuditQuestionRead, AuditQuestionUpdate

router = APIRouter(prefix="/api/questions", tags=["questions"])

EXCEL_COLUMNS = [
    ("id", "題目ID"),
    ("cycle_name", "稽核循環"),
    ("question", "題目內容"),
    ("procedure", "查核程序"),
    ("risk_description", "風險說明"),
    ("regulation_reference", "法規依據"),
    ("department", "適用部門"),
    ("enabled", "啟用狀態"),
]

EXCEL_HEADER_ALIASES = {key: {key, label} for key, label in EXCEL_COLUMNS}


def parse_enabled(value) -> bool:
    if value in (None, ""):
        return True
    normalized = str(value).strip().lower()
    if normalized in {"true", "1", "yes", "y", "是", "啟用", "啟用中", "已啟用", "enabled", "active"}:
        return True
    if normalized in {"false", "0", "no", "n", "否", "停用", "停用中", "已停用", "disabled", "inactive"}:
        return False
    raise HTTPException(status_code=400, detail=f"Invalid enabled value: {value}")


def validate_template(header: list[str]) -> None:
    expected = [label for _, label in EXCEL_COLUMNS]
    keys = [key for key, _ in EXCEL_COLUMNS]
    for index, key in enumerate(keys):
        value = header[index] if index < len(header) else ""
        if value not in EXCEL_HEADER_ALIASES[key]:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Invalid template. Expected columns: {', '.join(expected)}. "
                    f"Got: {', '.join(header)}"
                ),
            )


@router.get("", response_model=ApiResponse)
def list_questions(
    keyword: str | None = None,
    cycle_name: str | None = None,
    department: str | None = None,
    enabled: bool | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(current_user),
):
    query = db.query(AuditQuestion)
    if keyword:
        like = f"%{keyword}%"
        query = query.filter(or_(AuditQuestion.question.like(like), AuditQuestion.procedure.like(like)))
    if cycle_name:
        query = query.filter(AuditQuestion.cycle_name == cycle_name)
    if department:
        query = query.filter(AuditQuestion.department == department)
    if enabled is not None:
        query = query.filter(AuditQuestion.enabled == enabled)
    rows = query.order_by(AuditQuestion.created_at.desc()).all()
    return ApiResponse(message="Data retrieved successfully", data=[AuditQuestionRead.model_validate(row) for row in rows])


@router.get("/export")
def export_questions(db: Session = Depends(get_db), _: User = Depends(current_user)):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "questions"
    sheet.append([label for _, label in EXCEL_COLUMNS])

    rows = db.query(AuditQuestion).order_by(AuditQuestion.cycle_name.asc(), AuditQuestion.created_at.desc()).all()
    for row in rows:
        sheet.append(
            [
                row.id,
                row.cycle_name,
                row.question,
                row.procedure,
                row.risk_description,
                row.regulation_reference or "",
                row.department,
                "是" if row.enabled else "否",
            ]
        )

    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells) + 4
        sheet.column_dimensions[column_cells[0].column_letter].width = min(width, 50)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"audit_questions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/import", response_model=ApiResponse)
async def import_questions(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(current_user),
):
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")

    content = await file.read()
    workbook = load_workbook(BytesIO(content), data_only=True)
    sheet = workbook.active
    header = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    validate_template(header)

    keys = [key for key, _ in EXCEL_COLUMNS]
    created_count = 0
    updated_count = 0
    for row_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(values):
            continue
        data = dict(zip(keys, values))
        payload = {
            "cycle_name": str(data.get("cycle_name") or "").strip(),
            "question": str(data.get("question") or "").strip(),
            "procedure": str(data.get("procedure") or "").strip(),
            "risk_description": str(data.get("risk_description") or "").strip(),
            "regulation_reference": str(data.get("regulation_reference") or "").strip() or None,
            "department": str(data.get("department") or "").strip(),
            "enabled": parse_enabled(data.get("enabled")),
        }
        missing = [key for key in ("cycle_name", "question", "procedure", "risk_description", "department") if not payload[key]]
        if missing:
            raise HTTPException(status_code=400, detail=f"Row {row_index}: missing required fields: {', '.join(missing)}")

        question_id = str(data.get("id") or "").strip()
        row = db.get(AuditQuestion, question_id) if question_id else None
        if row:
            for key, value in payload.items():
                setattr(row, key, value)
            updated_count += 1
        else:
            if question_id:
                payload["id"] = question_id
            db.add(AuditQuestion(**payload))
            created_count += 1

    db.commit()
    return ApiResponse(
        message="Questions imported",
        data={"created": created_count, "updated": updated_count},
    )


@router.post("", response_model=ApiResponse)
def create_question(payload: AuditQuestionCreate, db: Session = Depends(get_db), _: User = Depends(current_user)):
    row = AuditQuestion(**payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return ApiResponse(message="Question created", data=AuditQuestionRead.model_validate(row))


@router.put("/{question_id}", response_model=ApiResponse)
def update_question(
    question_id: str,
    payload: AuditQuestionUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(current_user),
):
    row = db.get(AuditQuestion, question_id)
    if not row:
        raise HTTPException(status_code=404, detail="Question not found")
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(row, key, value)
    db.commit()
    db.refresh(row)
    return ApiResponse(message="Question updated", data=AuditQuestionRead.model_validate(row))
